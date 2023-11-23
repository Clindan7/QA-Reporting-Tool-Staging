import os

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from project.models import Projects
from report.models import TestCases, Sweep
from report.models.test_cases import SweepBugs
from report.models.testcase_result import TestcaseResult
from report.models.summary import Summary
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

current_datetime = datetime.datetime.now()
TEST_CASE_AND_RESULT = 'Test Case & Result'


# method for  custom column height
def custom_row_height(worksheet, start_row, end_row, height):
    for row_idx in range(start_row, end_row):
        worksheet.row_dimensions[row_idx].height = height


# method for  custom column width
def custom_column_width(worksheet, start_col, end_col, width):
    for col_idx in range(start_col, end_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


#  method for customizing cells
def custom_cells(worksheet, cell_name, cell_value, cell_font, cell_alignment, cell_fill=None):
    cell = worksheet[cell_name]
    cell.value = cell_value
    cell.font = cell_font
    cell.alignment = cell_alignment
    if cell_fill is not None:
        cell.fill = cell_fill
    else:
        cell.fill = PatternFill(
            start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # Default white fill


# common method for adding border to specified cells
def add_border(worksheet, rows, columns, border_style='thin', color='000000', left=True, right=True, top=True,
               bottom=True):
    for row_idx in range(rows[0], rows[1] + 1):
        for col_idx in range(columns[0], columns[1] + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_border = Border(
                left=Side(border_style=border_style,
                          color=color) if left else None,
                right=Side(border_style=border_style,
                           color=color) if right else None,
                top=Side(border_style=border_style,
                         color=color) if top else None,
                bottom=Side(border_style=border_style,
                            color=color) if bottom else None
            )
            cell.border = cell_border


# common method for formatting headers
def format_headers(worksheet, headers, header_font, title_fill, center_alignment, logger, start_row, end_row,
                   start_column, end_column, start_index=2):
    for col_idx, header in enumerate(headers, start=start_index):
        header_cell = worksheet.cell(
            row=start_row, column=col_idx, value=header)
        header_cell.font = header_font
        header_cell.fill = title_fill
        header_cell.alignment = center_alignment
        add_border(worksheet, rows=(start_row, end_row),
                   columns=(start_column, end_column))

    # Adjust column widths to fit content
    auto_adjust_column(worksheet, start_row, end_row, start_column, logger)


def auto_adjust_column(worksheet, start_row, end_row, start_column, logger):
    for col in worksheet.iter_cols(min_row=start_row, max_row=end_row, min_col=start_column,
                                   max_col=worksheet.max_column):
        max_length = 0
        for cell in col:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except ValueError as e:
                logger.error(
                    "ValueError occurred while processing auto-sizing cells: %s", str(e))

        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col[0].column_letter].width = adjusted_width


# common method for fetching data from db
def get_test_case_reports(model, logger, sheet_name=None, project_id=None, test_cases=None):
    data = model.objects.all()

    if project_id is not None:
        if sheet_name is not None:
            data = data.filter(sheet_name=sheet_name)
        data = data.filter(project=project_id)

    if test_cases:
        data = data.filter(test_cases=test_cases)

    if not data and sheet_name and project_id:
        logger.error(
            f"No data found in db for sheet_name '{sheet_name}' and project_id {project_id}")

    return data


def write_data_to_cells(worksheet, model_data, start_row, headers, alignment=None, start_column=2, max_characters=50):
    border_style = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000')
                          )

    if not alignment:
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable text wrapping

    last_row = None

    for row_idx, data in enumerate(model_data, start=start_row):
        for col_idx, header in enumerate(headers, start=start_column):
            # Get the corresponding attribute value from the model data
            value = getattr(data, header)

            # Check if the header is one of the specified headers and the length of the value exceeds max_characters
            if header in ["Test Cases Description", "Test Steps", "Pre Condition", "Expected Result"] and len(
                    value) > max_characters:
                alignment.wrap_text = True  # Enable text wrapping

            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border_style
            cell.alignment = alignment

        # Update last_row at the end of each iteration
        last_row = row_idx

    return last_row


# fetching all the tc sheet names and return it as a list
def fetch_all_testcase_sheets(project_id, logger):
    # fetch all testcase sheet under the project
    unique_sheets = set(
        TestCases.objects.filter(project=project_id)
        .values_list("sheet_name", flat=True)
        .distinct()
    )
    if unique_sheets:
        logger.info("Sheets exist: %s", unique_sheets)
    else:
        logger.error(
            "No module test case sheets exist for project", project_id)
    return unique_sheets


# method for fetching sweep data under a testcase and sweep count
def fetch_sweep_data(worksheet, all_test_cases, start_row, start_column, headers, sweep_count, alignment):
    data_ids = [item.id for item in all_test_cases]
    for data_id in data_ids:
        data = Sweep.objects.filter(
            test_cases=data_id, sweep_count=sweep_count)

        # for write data to actual result and build
        write_data_to_cells(worksheet, model_data=data, start_row=start_row,
                            headers=headers, start_column=start_column)

        # write data to sweep status column
        set_dropdown_statuses(worksheet, model_data=data, start_row=start_row, column=start_column + 2,
                              status_value='sweep_status', alignment=alignment)
        #  for fetching data from sweep bugs table
        for sweep in data:
            bug_ids = SweepBugs.objects.filter(
                sweep=sweep.id)

            bug_id_list = [bug.bug_id for bug in bug_ids]
            bug_id_string = ", ".join(map(str, bug_id_list))
            sweep.bug_id_list = bug_id_string

        # for write data to bug id
        db_bug_id_headers = ['bug_id_list']
        write_data_to_cells(worksheet, model_data=data, start_row=start_row,
                            headers=db_bug_id_headers, start_column=start_column + 3)

        start_row += 1


# method for creating each sweep headers
def sweeps(cell_range, merge_row, cell_name, cell_value, worksheet, header_font, center_alignment, title_fill, logger,
           start_row=7, end_row=7, start_column=15, end_column=19):
    worksheet.merge_cells(cell_range)
    add_border(worksheet, rows=(merge_row, merge_row),
               columns=(start_column, end_column))
    # add_border(worksheet, rows=(start_row, end_row), columns=(start_column, end_column))
    custom_cells(worksheet, cell_name, cell_value,
                 header_font, center_alignment, title_fill)
    sweep_headers = ['Actual Result', 'Build', 'Pass/ Fail', 'Bug ID']
    format_headers(worksheet, sweep_headers, header_font, title_fill, center_alignment, logger, start_row=start_row,
                   end_row=end_row, start_column=start_column, end_column=end_column, start_index=start_column)


#  define large column width
def common_large_columns(worksheet):
    # test case description
    custom_column_width(worksheet, start_col=4, end_col=4, width=45)
    # test steps
    custom_column_width(worksheet, start_col=7, end_col=7, width=45)
    # pre condition
    custom_column_width(worksheet, start_col=8, end_col=8, width=30)
    # expected result
    custom_column_width(worksheet, start_col=9, end_col=9, width=45)
    # actual result
    custom_column_width(worksheet, start_col=12, end_col=12, width=45)
    # actual result
    custom_column_width(worksheet, start_col=16, end_col=16, width=45)
    # actual result
    custom_column_width(worksheet, start_col=20, end_col=20, width=45)


# create dropdowns in specified rows and columns
def create_dropdown(worksheet, start_row, end_row, column, dropdown_values, center_alignment):
    # Define the DataValidation object
    data_validation = DataValidation(
        type="list", formula1=f'"{",".join(dropdown_values)}"')
    # Apply DataValidation to the specified range
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.alignment = center_alignment
        worksheet.add_data_validation(data_validation)
        data_validation.add(cell)


# set value to the created drop-down by fetching from db
def set_dropdown_statuses(worksheet, model_data, start_row, column, status_value, alignment):
    dropdown_values = ['pass', 'fail', 'untested']
    # make a list status in all_test_issues
    if status_value == 'status':
        status_data = [item.status for item in model_data]
    if status_value == 'sweep_status':
        status_data = [item.sweep_status for item in model_data]
    #  find the last column to be inserted
    end_row = start_row + len(status_data) - 1
    create_dropdown(worksheet, start_row, end_row,
                    column, dropdown_values, alignment)

    for row, status in enumerate(status_data, start=start_row):
        cell = worksheet.cell(row=row, column=column)
        cell.value = status
        add_border(worksheet, rows=(start_row, row), columns=(column, column))
        row += 1


# cover and version sheet
def sheet_1(worksheet, project_id, logger):
    logger.info("started cover and version of tc sheet")
    testcases = get_test_case_reports(
        TestcaseResult, logger, sheet_name='Cover & Version', project_id=project_id, )
    worksheet.title = 'Cover & Version'
    header_font = Font(bold=True, color="000000")
    main_header_font = Font(bold=True, size=18, color="000000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    title_fill = PatternFill(start_color="FF8EAADB",
                             end_color="FF8EAADB", fill_type="solid")

    custom_row_height(worksheet, 1, 4, 30)
    worksheet.merge_cells('A1:E3')
    custom_cells(worksheet, "A1", 'Innovature Software Labs P Ltd',
                 header_font, center_alignment)

    worksheet.merge_cells('F1:G3')
    img_url = os.path.join(os.path.dirname(__file__), 'images', 'inv-logo.png')
    # Create the Image object using the resolved URL
    img = Image(img_url)
    img.width = img.width * 1.35
    img.height = img.height * 1.25
    worksheet.add_image(img, 'F1')
    image_worksheet = worksheet['F1']
    image_worksheet.border = Border(
        right=Side(border_style='thin', color='000000'))
    add_border(worksheet, rows=(1, 3), columns=(7, 7), right=True)

    #  row 4 and 5
    worksheet.merge_cells('A4:B5')
    custom_cells(worksheet, "A4", 'ISL-ISMS-50-Test Cases & results',
                 header_font, center_alignment)
    add_border(worksheet, rows=(4, 5), columns=(1, 2))

    worksheet.merge_cells('C4:F5')
    custom_cells(worksheet, "C4", 'Test Cases & results',
                 header_font, center_alignment)
    add_border(worksheet, rows=(4, 5), columns=(3, 5))

    worksheet.merge_cells('G4:G5')
    custom_cells(worksheet, "G4", ' Revision Date: 17-07-2023',
                 header_font, center_alignment)
    add_border(worksheet, rows=(4, 5), columns=(6, 7))

    # Merge cells and set title with center alignment
    worksheet.merge_cells('B9:G11')
    custom_cells(worksheet, "B9", 'Test Case & Result ',
                 main_header_font, center_alignment, title_fill)
    add_border(worksheet, rows=(9, 11), columns=(2, 7), border_style='medium')

    worksheet.merge_cells('B15:G15')
    custom_cells(worksheet, "B15", 'Amendment Table of Test Case & Result Template',
                 header_font, center_alignment, title_fill)
    add_border(worksheet, rows=(15, 15), columns=(2, 7))

    headers = ['Version No.', 'Date of Release', 'Prepared by',
               'Reviewed by', 'Approved by', 'Change Description']
    header_font = Font(bold=True)
    format_headers(worksheet, headers, header_font, title_fill, center_alignment, logger, start_row=16, end_row=16,
                   start_column=2, end_column=7)

    # getting data from testcase result table
    db_headers = ['version', 'date_of_release', 'prepared_by',
                  'reviewed_by', 'approved_by', 'change_description']
    write_data_to_cells(worksheet, testcases, start_row=17, headers=db_headers)
    logger.info("completed cover and version of tc sheet")


# Amendment of tc sheet

def sheet_2(workbook, project_id, logger):
    logger.info("started Amendment of TC sheet")
    worksheet = workbook.create_sheet('Amendment of TC')
    worksheet.sheet_view.showGridLines = False

    main_header_font = Font(bold=True, size=18, color="000000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    title_fill = PatternFill(start_color="FF8EAADB",
                             end_color="FF8EAADB", fill_type="solid")

    worksheet.merge_cells('B2:G2')
    custom_row_height(worksheet, 2, 3, 30)
    custom_cells(worksheet, "B2", TEST_CASE_AND_RESULT,
                 main_header_font, center_alignment, title_fill)
    add_border(worksheet, rows=(2, 2), columns=(2, 7), border_style='medium')

    headers = ['Version No.', 'Date of Release', 'Prepared by',
               'Reviewed by', 'Approved by', 'Change Description']
    header_font = Font(bold=True)
    format_headers(worksheet, headers, header_font, title_fill, center_alignment, logger, start_row=4, end_row=4,
                   start_column=2, end_column=7)

    # getting data from testcase result table
    testcases = get_test_case_reports(
        TestcaseResult, logger, sheet_name='Amendment of TC', project_id=project_id)
    db_headers = ['version', 'date_of_release', 'prepared_by',
                  'reviewed_by', 'approved_by', 'change_description']
    write_data_to_cells(worksheet, testcases, start_row=5, headers=db_headers)
    logger.info("completed amendment of tc sheet")


# summary sheet
def sheet_3(workbook, project_id, logger):
    logger.info("started summary sheet")
    worksheet = workbook.create_sheet('Summary')
    worksheet.sheet_view.showGridLines = False

    header_font = Font(bold=True, size=10, color="000000")
    normal_font = Font(bold=False, size=10, color="000000")
    main_header_font = Font(bold=True, size=18, color="000000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    title_fill = PatternFill(start_color="FF8EAADB",
                             end_color="FF8EAADB", fill_type="solid")

    worksheet.merge_cells('B2:G2')
    custom_row_height(worksheet, 2, 3, 30)
    custom_cells(worksheet, "B2", TEST_CASE_AND_RESULT,
                 main_header_font, center_alignment, title_fill)
    add_border(worksheet, rows=(2, 2), columns=(2, 7), border_style='medium')

    # row 5
    custom_row_height(worksheet, 4, 5, 25)
    custom_cells(worksheet, "B4", 'Project Name ',
                 header_font, center_alignment, title_fill)
    worksheet.merge_cells('C4:D4')
    data = Projects.objects.filter(id=project_id)
    project_name = [item.name for item in data]
    custom_cells(worksheet, "C4",
                 project_name[0], normal_font, center_alignment)
    add_border(worksheet, rows=(4, 4), columns=(2, 4))

    # header row
    headers = ['Module/ Feature', 'No. of Test Cases', 'No. of Test Cases Executed',
               'No. of Test Cases Passed', 'No. of Test Cases Failed', 'No. of Test Cases Not tested']
    headers_font = Font(bold=True)
    custom_row_height(worksheet, 7, 8, 25)
    format_headers(worksheet, headers, headers_font, title_fill, center_alignment, logger, start_row=7, end_row=7,
                   start_column=2, end_column=7)
    custom_column_width(worksheet, start_col=3, end_col=7, width=28)
    custom_column_width(worksheet, start_col=2, end_col=2, width=40)

    # getting data from testcase result table
    summary = get_test_case_reports(Summary, logger, project_id=project_id)
    db_headers = ['feature', 'number_of_testcases', 'executed_testcases_count',
                  'passed_testcases_count', 'failed_testcases_count', 'not_tested_count']
    write_data_to_cells(worksheet, summary, start_row=8, headers=db_headers)
    logger.info("completed summary sheet")
    return project_name[0]


#  process each tc sheet
def module_test_case_sheet(workbook, project_id, sheet, logger):
    logger.info(f"started module unit tc sheet named : {sheet}")
    worksheet = workbook.create_sheet(sheet)
    worksheet.sheet_view.showGridLines = False
    header_font = Font(bold=True, color="000000")
    normal_font = Font(bold=False, size=10, color="000000")
    main_header_font = Font(bold=True, size=18, color="000000")
    center_alignment = Alignment(horizontal='center', vertical='center')
    title_fill = PatternFill(start_color="FF8EAADB",
                             end_color="FF8EAADB", fill_type="solid")

    # sheet title row
    worksheet.merge_cells('B1:X1')
    custom_row_height(worksheet, 1, 2, 30)
    custom_cells(worksheet, "B1", TEST_CASE_AND_RESULT,
                 main_header_font, center_alignment, title_fill)
    add_border(worksheet, rows=(1, 1), columns=(2, 24), border_style='medium')

    # row 5
    worksheet.merge_cells('B3:C3')
    custom_row_height(worksheet, 3, 4, 25)
    custom_cells(worksheet, "B3", 'Test Case Category ',
                 header_font, center_alignment, title_fill)
    custom_column_width(worksheet, start_col=3, end_col=4, width=65)
    test_case_category = 'API Testing'
    custom_cells(worksheet, "D3", test_case_category,
                 normal_font, center_alignment)
    add_border(worksheet, rows=(3, 3), columns=(2, 4))
    #  headers of module test cases
    # a) -----------------header rows from column B to J-------------------------#
    worksheet.merge_cells('B5:B6')
    worksheet.merge_cells('C5:C6')
    worksheet.merge_cells('D5:D6')
    worksheet.merge_cells('E5:E6')
    worksheet.merge_cells('F5:F6')
    worksheet.merge_cells('G5:G6')
    worksheet.merge_cells('H5:H6')
    worksheet.merge_cells('I5:I6')
    worksheet.merge_cells('J5:J6')
    worksheet.merge_cells('K5:K6')
    headers = ['Sl.No', 'TestCase \n ID', 'Test Cases \n Description', 'Feature', 'Sub Feature',
               'Test Steps', 'Pre Condition', 'Expected Result', 'Category', 'Status']
    headers_font = Font(bold=True)
    custom_row_height(worksheet, 5, 6, 25)
    format_headers(worksheet, headers, headers_font, title_fill, center_alignment, logger, start_row=5, end_row=6,
                   start_column=2, end_column=11)
    custom_column_width(worksheet, start_col=2, end_col=11, width=18)

    #  fetch data from db having below credentials
    all_test_cases = get_test_case_reports(
        TestCases, logger, sheet_name=sheet, project_id=project_id)
    #  set Sl No to the list obtained from db
    for index, item in enumerate(all_test_cases, start=1):
        item.sl_no = index

    db_headers = ['sl_no', 'test_case_id', 'description', 'feature', 'sub_feature', 'test_steps',
                  'pre_condition', 'expected_results', 'category']
    # write data to table until category column
    write_data_to_cells(worksheet, all_test_cases, start_row=7,
                        headers=db_headers, start_column=2)
    # write data to status column
    set_dropdown_statuses(worksheet, model_data=all_test_cases, start_row=7, column=11, status_value='status',
                          alignment=center_alignment)

    # B) -----------------header row for sweeps-------------------------#
    # for sweep 1
    sweeps(cell_range='L5:O5', merge_row=5, cell_name="L5", cell_value='Sweep 1', worksheet=worksheet,
           header_font=header_font, center_alignment=center_alignment, title_fill=title_fill, logger=logger,
           start_row=6, end_row=6, start_column=12, end_column=15
           )
    db_sweep_headers = ['actual_result', 'build']
    # 'sweep_status', 'bug_id']
    fetch_sweep_data(worksheet, all_test_cases, start_row=7, start_column=12, headers=db_sweep_headers,
                     sweep_count='1', alignment=center_alignment)

    # for sweep 2
    sweeps(cell_range='P5:S5', merge_row=5, cell_name="P5", cell_value='Sweep 2', worksheet=worksheet,
           header_font=header_font, center_alignment=center_alignment, title_fill=title_fill, logger=logger,
           start_row=6, end_row=6, start_column=16, end_column=19
           )
    fetch_sweep_data(worksheet, all_test_cases, start_row=7, start_column=16, headers=db_sweep_headers,
                     sweep_count='2', alignment=center_alignment)

    # for sweep 3
    sweeps(cell_range='T5:W5', merge_row=5, cell_name="T5", cell_value='Sweep 3', worksheet=worksheet,
           header_font=header_font, center_alignment=center_alignment, title_fill=title_fill, logger=logger,
           start_row=6, end_row=6, start_column=20, end_column=23
           )
    fetch_sweep_data(worksheet, all_test_cases, start_row=7, start_column=20, headers=db_sweep_headers,
                     sweep_count='3', alignment=center_alignment)

    # C) -----------------header row for comments-------------------------#
    worksheet.merge_cells('X5:X6')
    headers_comments = ['Comments']
    format_headers(worksheet, headers_comments, headers_font, title_fill, center_alignment, logger, start_row=5,
                   end_row=6, start_column=24, end_column=24, start_index=24)
    custom_column_width(worksheet, start_col=24, end_col=24, width=18)
    db_headers = ['comments']
    # write data to table
    last_row = write_data_to_cells(worksheet, all_test_cases, start_row=7,
                                   headers=db_headers, start_column=24)
    custom_row_height(worksheet, 7, last_row + 1, 130)
    common_large_columns(worksheet)
    logger.info(f"completed writing into sheet named {sheet}")


def export_to_excel(project_id, logger):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.sheet_view.showGridLines = False

    #  sheet 1 to 3 are common for all tc
    sheet_1(worksheet, project_id, logger)
    sheet_2(workbook, project_id, logger)
    project_name = sheet_3(workbook, project_id, logger)

    # fetch remaining tc sheet names
    logger.info("started fetching of all sheet names of module unit tc in db")
    all_sheets_under_project = fetch_all_testcase_sheets(project_id, logger)

    # process each tc sheets
    for sheet in all_sheets_under_project:
        logger.info(f"started processing of the sheet named : {sheet}")
        module_test_case_sheet(workbook, project_id, sheet, logger)

    logger.info(
        f"all sheets under the project {project_name} are created successfully")

    # file configurations
    export_dir = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), 'exported_reports')
    os.makedirs(export_dir, exist_ok=True)  # Ensure the directory exists
    file_name = f'ISL-ISMS-50-Test_Case_Result_{project_name}_Unit_TC.xlsx'
    excel_file_path = os.path.join(export_dir, file_name)
    # Save the Excel file
    workbook.save(excel_file_path)
    logger.info(f"generated excel file location ->{excel_file_path}")

    return file_name
